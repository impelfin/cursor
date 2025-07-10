import os
import re
from flask import Flask, request, render_template, send_file, url_for
import openpyxl
import pandas as pd
from io import BytesIO
from openpyxl.utils.cell import range_boundaries
import requests
from bs4 import BeautifulSoup
import time
import traceback 

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

EPOST_SEARCH_URL = "https://www.epost.go.kr/search.RetrieveIntegrationNewZipCdList.comm"

province_city_mapping = {
    '서울특별시': '서울특별시', '서울시': '서울특별시', '서울': '서울특별시',
    '부산광역시': '부산광역시', '부산시': '부산광역시', '부산': '부산광역시',
    '대구광역시': '대구광역시', '대구시': '대구광역시', '대구': '대구광역시',
    '인천광역시': '인천광역시', '인천시': '인천광역시', '인천': '인천광역시',
    '광주광역시': '광주광역시', '광주시': '광주광역시', '광주': '광주광역시', 
    '대전광역시': '대전광역시', '대전시': '대전광역시', '대전': '대전광역시',
    '울산광역시': '울산광역시', '울산시': '울산광역시', '울산': '울산광역시',
    '세종특별자치시': '세종특별자치시', '세종시': '세종특별자치시', '세종': '세종특별자치시',
    '제주특별자치도': '제주특별자치도', '제주도': '제주특별자치도', '제주': '제주특별자치도',
    '전라남도': '전라남도', '전남': '전라남도',
    '전라북도': '전라북도', '전북': '전라북도',
    '경상남도': '경상남도', '경남': '경상남도',
    '경상북도': '경상북도', '경북': '경상북도',
    '충청남도': '충청남도', '충남': '충청남도',
    '충청북도': '충청북도', '충북': '충청북도',
    '강원특별자치도': '강원특별자치도', '강원도': '강원특별자치도', '강원': '강원특별자치도',
    '경기도': '경기도', '경기': '경기도'
}

def clean_address_for_search(address):
    """
    주소 문자열을 검색에 적합한 형태로 정규화합니다.
    (아파트 동호수, 괄호 안 내용, 불필요한 특수문자 제거, 시/도 명칭 통일)
    """
    address = str(address).strip()
    
    # --- 1단계: 주소 상세 정보 제거 ---
    # 괄호 안 내용 제거
    address = re.sub(r'\s*\(.*?\)', '', address)
    # 동호수 제거 (예: 101동 102호, 102호)
    address = re.sub(r'\s*\d+동\s*\d+호|\s*\d+호', '', address)
    # '번지' 제거
    address = address.replace('번지', '')
    # 도로명 주소 형식 정규화 (예: '번길10' -> '번길 10', '로10-1' -> '로 10-1')
    address = re.sub(r'번길(\d+)', r'번길 \1', address)
    address = re.sub(r'(로|길)(\d+)-(\d+)', r'\1 \2-\3', address)
    address = re.sub(r'(로|길)(\d+)', r'\1 \2', address)
    
    # 연속된 공백을 하나로 줄임
    address = re.sub(r'\s+', ' ', address).strip()

    # --- 2단계: 시/도 명칭 표준화 ---
    matched_region_processed = False
    for old_name in sorted(province_city_mapping.keys(), key=len, reverse=True): 
        if address.startswith(old_name):
            target_full_name = province_city_mapping[old_name]
            
            if old_name == target_full_name:
                matched_region_processed = True
                break 
            
            address = address.replace(old_name, target_full_name, 1)
            matched_region_processed = True
            break 
    
    # --- 3단계: '시'로 끝나는 주소에 '도' 정보 추가 (예: 목포시 -> 전라남도 목포시) ---
    if not matched_region_processed and address.endswith('시'):
        jeonnam_cities_only_name = ['목포시', '여수시', '순천시', '나주시', '광양시'] 
        address_first_word = address.split(' ')[0]
        if address_first_word in jeonnam_cities_only_name and not address.startswith('전라남도'):
            address = '전라남도 ' + address
            
    address = re.sub(r'\s+', ' ', address).strip()
    
    return address

def _perform_epost_crawl(search_keyword):
    """
    실제로 우체국 웹사이트에 요청을 보내고 우편번호를 파싱하는 내부 함수.
    성공 시 (우편번호, 검색에 사용된 키워드) 튜플 반환. 실패 시 (None, None) 또는 (오류코드, None) 반환.
    """
    if not search_keyword or len(search_keyword) < 5:
        return None, None # 유효하지 않은 키워드
    
    try:
        payload = {
            'addr': search_keyword, 
            'gbn': 'postnew', 
            'targetDt': '', 
            'currentPage': '1',
            'countPerPage': '10', 
            'keyword': search_keyword 
        }
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://www.epost.go.kr/search.RetrieveIntegrationNewZipCdList.comm', 
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive'
        }

        response = requests.post(EPOST_SEARCH_URL, data=payload, headers=headers, timeout=10)
        response.raise_for_status() 

        soup = BeautifulSoup(response.text, 'html.parser')
        
        table = soup.find('table', class_='table_list')
        if not table:
            return None, None

        tbody = table.find('tbody')
        if not tbody:
            tbody = table 
        
        result_rows = tbody.find_all('tr')
        
        if not result_rows:
            return None, None

        for row in result_rows:
            postcode_th = row.find('th') 
            
            if postcode_th:
                postcode_text = postcode_th.get_text().strip()
                postcode_match = re.search(r'\b\d{5}\b', postcode_text)
                if postcode_match:
                    return postcode_match.group(0), search_keyword # 성공 시 우편번호와 검색 키워드 반환
        
        return None, None # 유효한 우편번호를 찾지 못함

    except requests.exceptions.RequestException as e:
        print(f"!!! HTTP 요청 오류 발생: {e} !!!")
        return "네트워크오류", None
    except Exception as e:
        print(f"!!! 예상치 못한 파싱 오류 발생: {e} !!!")
        traceback.print_exc()
        return "크롤링오류", None
    finally:
        time.sleep(0.5) # 요청 간 지연

def search_postcode_via_epost_crawl(original_full_address):
    original_full_address = str(original_full_address).strip()
    if not original_full_address:
        print(f"우체국 크롤링 시도: '{original_full_address}' -> 검색 키워드가 비어있습니다.")
        return None, None # 우편번호, 검색된 주소 반환

    # 주소 정규화 (괄호, 동호수 등 제거)
    base_cleaned_address = clean_address_for_search(original_full_address)
    
    # 시도명으로 시작하는 주소는 최대한 유지
    address_parts = base_cleaned_address.split(' ')
    
    # 시도명은 항상 포함되도록 시작 인덱스 설정
    start_index = 0
    if len(address_parts) > 1 and address_parts[0] in province_city_mapping.values():
        start_index = 1 # 시도명 다음부터 줄이기 시작

    # 검색 시도할 주소 목록 생성 (가장 긴 주소부터 짧은 주소 순)
    addresses_to_try = []
    # 원본 정규화 주소도 가장 먼저 시도
    if len(base_cleaned_address) >= 5:
        addresses_to_try.append(base_cleaned_address)

    # 한 단어씩 줄여가며 주소 생성
    for i in range(len(address_parts) - 1, start_index, -1): # 마지막 단어부터 줄이기 시작
        truncated_address = ' '.join(address_parts[:i])
        if len(truncated_address) >= 5: # 최소 길이 5자 이상 유지
            addresses_to_try.append(truncated_address)
    
    # 중복 제거 및 순서 유지
    seen = set()
    unique_addresses = []
    for addr in addresses_to_try:
        if addr not in seen:
            unique_addresses.append(addr)
            seen.add(addr)
    
    print(f"\n--- '{original_full_address}' 주소에 대해 여러 검색 시도 ---")
    for attempt_address in unique_addresses:
        print(f"  시도 중 주소: '{attempt_address}'")
        postcode, found_address = _perform_epost_crawl(attempt_address)
        
        if postcode and postcode not in ["네트워크오류", "크롤링오류"]:
            print(f"  성공: '{original_full_address}' -> 우편번호: {postcode} (검색 주소: '{found_address}')")
            return postcode, found_address # 우편번호와 검색된 주소 반환
        elif postcode == "네트워크오류" or postcode == "크롤링오류":
            # 네트워크 오류나 크롤링 오류는 재시도해도 의미 없으므로 바로 반환
            print(f"  오류 발생: {postcode} (검색 주소: '{attempt_address}')")
            return postcode, None # 오류 코드와 None 반환

    print(f"--- '{original_full_address}' 주소에 대한 모든 검색 시도 실패 ---")
    return "검색실패", None # 모든 시도 실패 시

@app.route('/', methods=['GET', 'POST'])
def index():
    download_url = None
    message = None

    if request.method == 'POST':
        if 'file' not in request.files:
            message = "파일이 없습니다."
            return render_template('index.html', message=message)
            
        file = request.files['file']
        if file.filename == '':
            message = "파일을 선택해주세요."
            return render_template('index.html', message=message)
        
        if file:
            try:
                in_memory_file = BytesIO()
                file.save(in_memory_file)
                in_memory_file.seek(0)

                workbook = openpyxl.load_workbook(in_memory_file)

                if 'Data' not in workbook.sheetnames:
                    message = "엑셀 파일에 'Data'라는 이름의 시트가 없습니다."
                    return render_template('index.html', message=message)
                
                sheet = workbook['Data']

                # 헤더 (5행) 수정
                # 3번째 열 (C열) 제목을 '우편번호'로 변경
                sheet.cell(row=5, column=3).value = '우편번호'
                # 6번째 열 (F열) 제목을 '검색된 주소'로 변경
                sheet.cell(row=5, column=6).value = '검색된 주소'


                merged_ranges_coords = []
                for merged_range in sheet.merged_cells:
                    if isinstance(merged_range, str): 
                        min_col, min_row, max_col, max_row = range_boundaries(merged_range) 
                    else: 
                        min_row = merged_range.min_row
                        max_row = merged_range.max_row
                        min_col = merged_range.min_col
                        max_col = merged_range.max_col
                    merged_ranges_coords.append((min_row, min_col, max_row, max_col))
                
                processed_count = 0
                total_rows = sheet.max_row - 1 
                
                print(f"\n--- 엑셀 파일 처리 시작 (총 {total_rows}개 주소 예상, 우체국 크롤링 사용) ---")
                
                for row_index in range(2, sheet.max_row + 1):
                    address_cell = sheet.cell(row=row_index, column=4)
                    address = address_cell.value

                    postcode_target_cell = sheet.cell(row=row_index, column=3)
                    found_address_target_cell = sheet.cell(row=row_index, column=6) # F열 (6번째 컬럼)

                    # 병합된 셀 확인
                    is_postcode_merged = False
                    is_found_address_merged = False
                    
                    for min_r, min_c, max_r, max_c in merged_ranges_coords:
                        if min_r <= postcode_target_cell.row <= max_r and min_c <= postcode_target_cell.column <= max_c:
                            is_postcode_merged = True
                        if min_r <= found_address_target_cell.row <= max_r and min_c <= found_address_target_cell.column <= max_c:
                            is_found_address_merged = True
                        if is_postcode_merged and is_found_address_merged:
                            break

                    if is_postcode_merged or is_found_address_merged:
                        print(f"경고: {postcode_target_cell.coordinate} 또는 {found_address_target_cell.coordinate} 셀이 병합된 셀의 일부입니다. 건너뜁니다.")
                        continue

                    if address:
                        postcode, found_address = search_postcode_via_epost_crawl(str(address)) 
                        
                        if postcode and postcode not in ["검색실패", "키워드짧음", "네트워크오류", "크롤링오류"]:
                            postcode_target_cell.value = postcode
                            found_address_target_cell.value = found_address # F열에 검색된 주소 추가
                            processed_count += 1
                        else:
                            postcode_target_cell.value = postcode 
                            found_address_target_cell.value = "" # 실패 시 F열 비움
                    else:
                        postcode_target_cell.value = ""
                        found_address_target_cell.value = "" # 주소 자체가 없으면 F열 비움
                    
                    if (row_index - 1) % 10 == 0 or (row_index - 1) == total_rows:
                        print(f"처리 중: {row_index - 1}/{total_rows} 행 완료...")

                processed_excel = BytesIO()
                workbook.save(processed_excel)
                processed_excel.seek(0)

                output_filename = f"processed_{file.filename}"
                output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                with open(output_filepath, 'wb') as f:
                    f.write(processed_excel.getvalue())
                
                download_url = url_for('download_file', filename=output_filename)
                message = f"파일 처리가 완료되었습니다! 총 {processed_count}개의 우편번호를 채웠습니다."
                print(f"\n--- 엑셀 파일 처리 완료! 총 {processed_count}개의 우편번호 채움 ---")

            except Exception as e:
                print(f"!!! 파일 처리 중 예외 발생: {e} !!!")
                message = f"파일 처리 중 오류 발생: {e}"

    return render_template('index.html', download_url=download_url, message=message)

@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        return "파일을 찾을 수 없습니다.", 404

if __name__ == '__main__':
    app.run(debug=True)
